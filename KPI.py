# Import the Outage File
import pandas as pd
import numpy as np
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# load excel file by location
for folders, sub_folders, file in os.walk('D:\\MSU\\KPIs\\'):
          for name in file:
                if (name.startswith("Out")):
                    filename = os.path.join(folders, name)
                    df = pd.read_excel(filename, usecols=[0,3,5,10,15,16,17,18,19,38])
                    break


# Removing Zagazig and Fayoum, also sorting by Duration
df = df.loc[(df['Region'] != "Menoufia") & (df['Region'] != "Zagazig") & (df['Region'] != "Fayoum")].sort_values(['Duration'], ascending=False)

regions = ['Delta',
 'Tagamo3',
 'Giza',
 'Downtown',
 'Gesr El Suiz',
 'Mohandesen',
 'Helwan']


## Extract insufficient comments and Owner

df['SuffComment'] =np.nan
df['Owner'] = np.nan
df['Zone'] = np.nan

df.loc[df['Region'].isin(["Helwan", "Downtown", "Giza"]), 'Zone'] = 'Z1'
df.loc[~df['Region'].isin(["Helwan", "Downtown", "Giza"]), 'Zone'] = 'Z2'

df.loc[df['Duration'] <= 60, 'Owner'] = 'FO'
df.loc[df['Duration'] > 60, 'Owner'] = 'FM'
df.loc[(df['SubCategory'].str.contains("fuel|gen", flags=re.I, regex=True)) & (df['Duration'] > 60), 'Owner'] = 'EM'
df['SuffComment'] = "OK"
df.loc[(df['Owner'] != "FO") & (df['RootCause'].isnull()), 'SuffComment'] = "NOT_OK"
df.loc[(df['Owner'] == "EM") & (df['Access'] != True) & (df['SubCategory'] == "Generator") &(df['RootCause'].str.contains("fail|down|cut|gen problem|generator problem", flags=re.I, regex=True)), 'SuffComment'] = "NOT_OK"
df.loc[(df['Owner'] != 'FO') & (df['RootCause'] == '\"\"'), 'SuffComment'] = "NOT_OK"
df.to_excel('modified.xlsx', index=False)


## Coloring the NOT_OK rows
df = pd.read_excel('modified.xlsx')
df = df.sort_values(['SuffComment', 'Owner'], ascending=[True, True])


def highlight(s):
    if s.SuffComment == "NOT_OK":
        if s.Owner == "EM":
            return ['background-color: #9370DB']*13
        elif s.Owner == "FM":
            return ['background-color: #BC8F8F']*13
    else:
        return ['background-color: transparent']*13
    
    
df = df.style.apply(highlight, axis=1)
df.to_excel('modified.xlsx', index=False)


## Getting Todays KPI's Sheet
df = pd.read_excel('modified.xlsx')
df = df.loc[df['SuffComment'] == "NOT_OK"]
df = df.style.apply(highlight, axis=1)
df.to_excel('Not_Ok.xlsx', index=False)


df = pd.read_excel('modified.xlsx')
df_fm = df.loc[df['Owner'] == "FM"]
df_em = df.loc[df['Owner'] == "EM"]

df_all_Pivot = pd.pivot_table(df,index=['Owner'],columns=['SuffComment'],values=['Site'],aggfunc='count')
df_all_Pivot = df_all_Pivot.fillna(0) #------>1

df_em_Pivot = pd.pivot_table(df_em,index=['Zone'],columns=['SuffComment'],values=['Site'],aggfunc='count')
df_em_Pivot = df_em_Pivot.fillna(0) #------>2

df_fm_Pivot = pd.pivot_table(df_fm,index=['Region'],columns=['SuffComment'],values=['Site'],aggfunc='count')
df_fm_Pivot = df_fm_Pivot.fillna(0) #------>3


## Getting Output
with pd.ExcelWriter('results.xlsx') as writer:
    df_all_Pivot.to_excel(writer, sheet_name='count_all')
    df_em_Pivot.to_excel(writer, sheet_name='EM_Zones')
    df_fm_Pivot.to_excel(writer, sheet_name='FM_Regions')
    
# df_all_part

df_all = pd.read_excel('results.xlsx', sheet_name='count_all')
df_all = df_all.drop([1])
df_all = df_all.reset_index()
df_all = df_all.drop(['index'], axis=1)
df_all.columns = ['Owner', 'NOT_OK', 'OK']
df_all = df_all.drop([0])

df_all['Total'] = df_all.iloc[:, 1:3].sum(axis=1)

# EM Part

df_em_Pivot = pd.read_excel('results.xlsx', sheet_name='EM_Zones')
df_em_Pivot = df_em_Pivot.drop([1])
df_em_Pivot = df_em_Pivot.reset_index()

df_em_Pivot = df_em_Pivot.drop(['index'], axis=1)
df_em_Pivot.columns = ['Owner', 'NOT_OK', 'OK']
df_em_Pivot = df_em_Pivot.drop([0])

df_em_Pivot['Total'] = df_em_Pivot.iloc[:, 1:3].sum(axis=1)

# FM Part

df_fm_Pivot = pd.read_excel('results.xlsx', sheet_name='FM_Regions')
df_fm_Pivot = df_fm_Pivot.drop([1])
df_fm_Pivot = df_fm_Pivot.reset_index()
df_fm_Pivot = df_fm_Pivot.drop(['index'], axis=1)

df_fm_Pivot.columns = ['Owner', 'NOT_OK', 'OK']

df_fm_Pivot = df_fm_Pivot.drop([0])

df_fm_Pivot['Total'] = df_fm_Pivot.iloc[:, 1:3].sum(axis=1)


with pd.ExcelWriter('results.xlsx') as writer:
    df_all.to_excel(writer, sheet_name='count_all')
    df_em_Pivot.to_excel(writer, sheet_name='EM_Zones')
    df_fm_Pivot.to_excel(writer, sheet_name='FM_Regions')
    
    
# Final Table
df1 = pd.read_excel('results.xlsx', sheet_name='FM_Regions')
df2 = pd.read_excel('results.xlsx', sheet_name='count_all')
df3 = pd.read_excel('results.xlsx', sheet_name='EM_Zones')

frames = [df1, df2, df3]

df_final = pd.concat(frames, sort=False).reset_index()
df_final = df_final.drop(['index'], axis=1)

######
#df_final['KPI'] = (df_final['OK'].div(df_final['Total'], axis=0)).round(4)
##########

df_final['index'] = [0,1,2,3,4,5,6,11,7,8,9,10]
df_final =df_final.sort_values(by=['index'], ascending=True)
df_final = df_final.drop(['index', 'Unnamed: 0'], axis=1)

####
#df_final['KPI'] *= 100
#df_final['KPI'] = (df_final['KPI'].round(2)).astype(str) + '%'
#####

df_final.to_excel('results.xlsx', index=False)


### New
wb1 = load_workbook('test.xlsx')
ws1 = wb1.active
wb2 = load_workbook('results.xlsx')
ws2 = wb2.active

for i in range(3, 15):
    ws1.cell(i,2).value = ws2.cell(i-1,2).value
    ws1.cell(i,3).value = ws2.cell(i-1,3).value
    ws1.cell(i,4).value = ws2.cell(i-1,4).value
    
wb1.save('results.xlsx')

print("Done!")
